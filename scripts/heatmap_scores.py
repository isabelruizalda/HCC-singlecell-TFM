"""
heatmap_scores.py
-----------------
Para cada grupo funcional (G1-G17), calcula el gene score de cada célula
usando sc.tl.score_genes. Luego, para cada tipo celular, calcula qué
PORCENTAJE DE CÉLULAS de ese tipo supera el umbral de score (percentil 75
calculado sobre TODAS las células del dataset).

Esto evita que tipos celulares muy abundantes (ej. Hepatocytes) dominen
el resultado por puro número, y es consistente con el heatmap de expresión.

Lee los genes de genes_filtrados_zscore.xlsx (generado por heatmap_permisivo.py)
para usar solo genes con expresión diferencial real (z-score >= 1).

Genera:
  - Un Excel con los porcentajes por grupo, tipo celular y dataset

Uso:
    python scripts/score_celltypes.py \\
        --data_dir data \\
        --genes_dir 170_genes \\
        --genes_file genes_filtrados_zscore.xlsx \\
        --figures_dir figures/score_celltypes \\
        --results_dir results \\
        --percentil 75
"""

import os
import glob
import argparse
import warnings
import numpy as np
import pandas as pd
import scanpy as sc
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from matplotlib.colors import Normalize
from matplotlib.cm import ScalarMappable
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")


# ──────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────────
CELLTYPES_ORDER = ["Fibroblast", "Endothelial", "Hepatocyte", "Myeloid", "T/NK", "B"]

CELLTYPE_COLORS = {
    "Fibroblast":  "#FFD54F",
    "Endothelial": "#81C784",
    "Hepatocyte":  "#FF6B6B",
    "Myeloid":     "#4FC3F7",
    "T/NK":        "#DA9EC4",
    "B":           "#FF8A65",
}

GENE_SYMBOL_CANDIDATES = ["feature_name", "gene_name", "gene_symbol", "Symbol", "symbol"]


def normalizar_var_names(adata, dataset_name):
    if not adata.var_names[0].startswith("ENSG"):
        return adata
    for col in GENE_SYMBOL_CANDIDATES:
        if col in adata.var.columns:
            adata.var_names = adata.var[col].astype(str).values
            adata.var_names_make_unique()
            adata.raw = None
            return adata
    return adata


# ──────────────────────────────────────────────────────────────
# CALCULAR % DE CÉLULAS POR TIPO CELULAR QUE ACTIVA LA FIRMA
# ──────────────────────────────────────────────────────────────
def calcular_porcentajes_por_celltype(adata, genes_grupo, score_key, percentil):
    """
    Opción B: calcula el score global, define el umbral como el percentil
    indicado sobre TODAS las células, y luego mide qué % de células de
    CADA TIPO CELULAR supera ese umbral.

    Esto evita el sesgo por abundancia celular.

    Devuelve:
        porcentajes: dict {celltype: % de células que superan el umbral}
        n_genes_presentes: int
    """
    adata_pp = adata.copy()

    # Normalizar solo si no está log-transformado
    if adata_pp.X.max() > 50:
        sc.pp.normalize_total(adata_pp, target_sum=1e4)
        sc.pp.log1p(adata_pp)

    # Genes presentes en el dataset
    genes_presentes = [g for g in genes_grupo if g in adata_pp.var_names]
    if len(genes_presentes) == 0:
        return None, 0

    # Calcular score para cada célula
    sc.tl.score_genes(adata_pp, gene_list=genes_presentes, score_name=score_key)
    scores = adata_pp.obs[score_key].values

    # Umbral global (percentil sobre todas las células)
    umbral = np.percentile(scores, percentil)

    # Para cada tipo celular: % de sus células que supera el umbral
    porcentajes = {}
    celltypes = adata_pp.obs["celltype"].values

    for ct in CELLTYPES_ORDER:
        mask_ct = celltypes == ct
        n_ct = mask_ct.sum()
        if n_ct == 0:
            continue
        n_activas = (scores[mask_ct] >= umbral).sum()
        porcentajes[ct] = round(n_activas / n_ct * 100, 1)

    # Otros tipos celulares no en CELLTYPES_ORDER
    mask_otros = np.array([ct not in CELLTYPES_ORDER for ct in celltypes])
    n_otros = mask_otros.sum()
    if n_otros > 0:
        n_activas_otros = (scores[mask_otros] >= umbral).sum()
        porcentajes["Otros"] = round(n_activas_otros / n_otros * 100, 1)

    return porcentajes, len(genes_presentes)




# ──────────────────────────────────────────────────────────────
# GUARDAR EXCEL
# ──────────────────────────────────────────────────────────────
def guardar_excel(todos_resultados, out_path, percentil):
    """
    Guarda un Excel con una hoja por dataset y una hoja resumen.
    Los valores son % de células del tipo celular que activan la firma.
    """
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ct_colors_hex = {
        "Fibroblast":  "FFD54F", "Endothelial": "81C784",
        "Hepatocyte":  "FF6B6B", "Myeloid":     "4FC3F7",
        "T/NK":        "DA9EC4", "B":            "FF8A65",
        "Otros":       "CCCCCC",
    }

    grupos = list(todos_resultados.keys())
    datasets = []
    for grupo, res in todos_resultados.items():
        for ds in res:
            if ds not in datasets:
                datasets.append(ds)

    # ── Hoja resumen (media entre datasets) ──
    ws = wb.create_sheet("Resumen (media datasets)")
    ws.cell(1, 1, f"% células con firma activa (umbral = percentil {percentil} global)").font = Font(bold=True, name="Arial", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CELLTYPES_ORDER) + 1)

    ws.cell(2, 1, "Grupo funcional").font = Font(bold=True, name="Arial", size=10)
    for j, ct in enumerate(CELLTYPES_ORDER, 2):
        c = ws.cell(2, j, ct)
        c.font = Font(bold=True, name="Arial", size=10, color="000000")
        c.fill = PatternFill("solid", fgColor=ct_colors_hex.get(ct, "FFFFFF"))
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for i, grupo in enumerate(grupos, 3):
        ws.cell(i, 1, grupo).font = Font(name="Arial", size=10)
        ws.cell(i, 1).border = border
        for j, ct in enumerate(CELLTYPES_ORDER, 2):
            vals = [res[ct] for ds, res in todos_resultados[grupo].items()
                    if res and ct in res]
            media = round(np.mean(vals), 1) if vals else 0
            c = ws.cell(i, j, media)
            c.number_format = '0.0"%"'
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center")
            c.border = border
            c.fill = PatternFill("solid", fgColor=ct_colors_hex.get(ct, "CCCCCC"))

    ws.column_dimensions["A"].width = 40
    for j in range(2, len(CELLTYPES_ORDER) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14

    # ── Una hoja por dataset ──
    for ds_id in datasets:
        ws2 = wb.create_sheet(ds_id[:31])
        ws2.cell(1, 1, f"% células con firma activa — {ds_id} — Percentil {percentil}").font = Font(bold=True, name="Arial", size=11)
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CELLTYPES_ORDER) + 1)

        ws2.cell(2, 1, "Grupo").font = Font(bold=True, name="Arial", size=10)
        for j, ct in enumerate(CELLTYPES_ORDER, 2):
            c = ws2.cell(2, j, ct)
            c.font = Font(bold=True, name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=ct_colors_hex.get(ct, "FFFFFF"))
            c.alignment = Alignment(horizontal="center")
            c.border = border

        for i, grupo in enumerate(grupos, 3):
            res = todos_resultados[grupo].get(ds_id)
            ws2.cell(i, 1, grupo).font = Font(name="Arial", size=10)
            ws2.cell(i, 1).border = border
            for j, ct in enumerate(CELLTYPES_ORDER, 2):
                val = res.get(ct, 0) if res else 0
                c = ws2.cell(i, j, val)
                c.number_format = '0.0"%"'
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center")
                c.border = border

        ws2.column_dimensions["A"].width = 40
        for j in range(2, len(CELLTYPES_ORDER) + 2):
            ws2.column_dimensions[get_column_letter(j)].width = 12

    wb.save(out_path)
    print(f"\n[OK] Excel guardado: {out_path}")



# ──────────────────────────────────────────────────────────────
# HEATMAP DE SCORES POR DATASET
# ──────────────────────────────────────────────────────────────
def generar_heatmap(todos_resultados, figures_dir, percentil):
    """
    Heatmap con filas=firmas, columnas agrupadas por tipo celular
    con una subcolumna por dataset.
    """
    grupos   = list(todos_resultados.keys())
    datasets = []
    for res in todos_resultados.values():
        for ds in res:
            if ds not in datasets:
                datasets.append(ds)

    n_grupos = len(grupos)
    n_ds     = len(datasets)
    n_cts    = len(CELLTYPES_ORDER)

    # Construir matriz: filas=grupos, cols=(ct x ds)
    matrix = np.zeros((n_grupos, n_cts * n_ds))
    for gi, grupo in enumerate(grupos):
        for ci, ct in enumerate(CELLTYPES_ORDER):
            for di, ds_id in enumerate(datasets):
                res = todos_resultados[grupo].get(ds_id)
                if res and ct in res:
                    matrix[gi, ci * n_ds + di] = res[ct]

    # Figura
    fig_w = max(20, n_cts * n_ds * 0.9 + 5)
    fig_h = max(10, n_grupos * 0.75 + 3)

    fig = plt.figure(figsize=(fig_w, fig_h))
    from matplotlib.gridspec import GridSpec
    gs  = fig.add_gridspec(1, 2, width_ratios=[fig_w - 0.8, 0.4], wspace=0.02)
    ax    = fig.add_subplot(gs[0])
    ax_cb = fig.add_subplot(gs[1])

    norm = Normalize(vmin=0, vmax=100)
    cmap = plt.get_cmap("RdBu_r")

    ax.imshow(matrix, aspect="auto", cmap=cmap, norm=norm, interpolation="nearest")

    # Eje Y — firmas
    ax.set_yticks(range(n_grupos))
    ax.set_yticklabels(grupos, fontsize=13, fontweight="bold")

    # Eje X inferior — datasets abreviados
    ds_short = [ds.replace("_HCC","").replace("_","·") for ds in datasets]
    col_labels = []
    for ct in CELLTYPES_ORDER:
        for ds in ds_short:
            col_labels.append(ds)
    ax.set_xticks(range(n_cts * n_ds))
    ax.set_xticklabels(col_labels, rotation=45, ha="right", fontsize=11, fontweight="bold")

    # Eje X superior — tipos celulares centrados
    ax2 = ax.twiny()
    ax2.set_xlim(ax.get_xlim())
    ct_positions = [ci * n_ds + (n_ds - 1) / 2 for ci in range(n_cts)]
    ax2.set_xticks(ct_positions)
    ax2.set_xticklabels(CELLTYPES_ORDER, fontsize=15, fontweight="bold")
    for tick, ct in zip(ax2.get_xticklabels(), CELLTYPES_ORDER):
        tick.set_color(CELLTYPE_COLORS.get(ct, "black"))

    # Líneas blancas separando tipos celulares
    for ci in range(1, n_cts):
        ax.axvline(ci * n_ds - 0.5, color="white", linewidth=2.5)

    # Números en celdas desactivados

    ax.set_title(
        f"% células con firma activa por dataset y tipo celular (percentil {percentil})",
        fontsize=13, fontweight="bold", pad=30)

    # Colorbar
    sm = ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = fig.colorbar(sm, cax=ax_cb)
    cbar.set_label("% células\ncon firma activa", fontsize=10)
    cbar.ax.tick_params(labelsize=9)

    # Leyenda datasets
    import matplotlib.patches as mpatches
    handles = [mpatches.Patch(color="gray", alpha=0.5, label=f"Col {i+1}: {ds}")
               for i, ds in enumerate(datasets)]
    fig.legend(handles=handles, fontsize=9, loc="lower center",
               ncol=n_ds, bbox_to_anchor=(0.45, -0.08),
               frameon=True, edgecolor="gray",
               title="Datasets", title_fontsize=9)

    plt.tight_layout()
    out = os.path.join(figures_dir, "heatmap_scores_datasets.png")
    fig.savefig(out, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close("all")
    print(f"\n[OK] Heatmap guardado: {out}")


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data_dir",    default="data")
    parser.add_argument("--genes_dir",   default="170_genes")
    parser.add_argument("--genes_file",  default="genes_filtrados_manual.xlsx")
    parser.add_argument("--figures_dir", default="figures/heatmap_scores")
    parser.add_argument("--results_dir", default="results")
    parser.add_argument("--percentil",   type=int, default=75)
    args = parser.parse_args()

    os.makedirs(args.figures_dir, exist_ok=True)
    os.makedirs(args.results_dir, exist_ok=True)

    # Cargar Excel de genes filtrados
    genes_path = os.path.join(args.genes_dir, args.genes_file)
    genes_df = pd.read_excel(genes_path)
    print(f"Cargando Excel: {genes_path}")
    print(f"Columnas: {genes_df.columns.tolist()}")

    gene_col  = "Gene"
    grupo_col = "Clasificacion_final"
    grupos = genes_df[grupo_col].dropna().unique()
    print(f"Grupos funcionales: {len(grupos)}")

    # Autodescubrir datasets
    h5ad_files = sorted(glob.glob(os.path.join(args.data_dir, "*.h5ad")))
    print(f"\nDatasets encontrados: {len(h5ad_files)}")

    # Cargar datasets
    datasets_loaded = {}
    for h5ad_path in h5ad_files:
        ds_id = os.path.basename(h5ad_path).replace(".h5ad", "").replace("adata_", "").replace("_raw", "")
        adata = sc.read_h5ad(h5ad_path)
        adata = normalizar_var_names(adata, ds_id)
        if "celltype" not in adata.obs.columns:
            print(f"  [SKIP] Sin celltype: {ds_id}")
            continue
        datasets_loaded[ds_id] = adata
        print(f"  Cargado {ds_id}: {adata.shape}")

    # Calcular porcentajes para cada grupo y dataset
    todos_resultados = {}

    for grupo in grupos:
        genes_grupo = genes_df[genes_df[grupo_col] == grupo][gene_col].dropna().tolist()
        score_key   = f"score_{grupo[:20].replace(' ', '_')}"

        print(f"\n{'='*60}")
        print(f"  Grupo: {grupo} ({len(genes_grupo)} genes tras filtro z-score)")
        print(f"{'='*60}")

        resultados_grupo = {}
        for ds_id, adata in datasets_loaded.items():
            print(f"  Calculando en {ds_id}...")
            porcentajes, n_genes = calcular_porcentajes_por_celltype(
                adata, genes_grupo, score_key, args.percentil
            )
            if porcentajes is None:
                print(f"    [SKIP] Sin genes del grupo en este dataset")
            else:
                print(f"    {n_genes} genes presentes — % células con firma activa:")
                for ct, pct in porcentajes.items():
                    if pct > 0:
                        print(f"      {ct}: {pct}%")
            resultados_grupo[ds_id] = porcentajes

        todos_resultados[grupo] = resultados_grupo


    # Guardar Excel
    excel_path = os.path.join(args.results_dir, f"score_celltypes_p{args.percentil}.xlsx")
    guardar_excel(todos_resultados, excel_path, args.percentil)

    # Heatmap de scores por dataset
    generar_heatmap(todos_resultados, args.figures_dir, args.percentil)

    print(f"\n{'='*60}")
    print(f"  Completado. Resultados en:")
    print(f"  - Figuras: {args.figures_dir}")
    print(f"  - Excel:   {excel_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
