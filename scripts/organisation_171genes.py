"""
=============================================================================
CLUSTERING JERÁRQUICO DE GENES — TFM Isabel Ruiz de Alda
=============================================================================

OBJETIVO:
    Clasificar 171 genes en grupos funcionales basándose en su perfil
    de anotación funcional (en qué vías aparecen).

INSPIRACIÓN METODOLÓGICA:
    Huang et al. (2007) DAVID Gene Functional Classification Tool.
    Genome Biology. https://pmc.ncbi.nlm.nih.gov/articles/PMC2375021/

    La clave del método de DAVID es:
    1. Construir una matriz binaria genes x anotaciones con TODAS las
       anotaciones disponibles (sin filtrar por FDR).
    2. Medir similitud entre genes basándose en cuántas anotaciones comparten.
    3. Agrupar genes con perfil similar mediante clustering.
    4. Caracterizar cada grupo con los términos más enriquecidos (aquí sí FDR).

POR QUÉ NO FILTRAMOS POR FDR PARA LA MATRIZ:
    El FDR mide si una VÍA está enriquecida en tu LISTA COMPLETA de genes.
    No mide si una anotación es válida para un gen individual.
    Usar solo vías FDR<0.05 excluye genes correctamente anotados que
    simplemente no tienen suficientes "compañeros" en tu lista.
    DAVID usa TODAS las anotaciones disponibles — nosotros hacemos lo mismo.

ARCHIVOS DE ENTRADA (en ../data/ relativo a scripts/):
    - Reactome_Pathways_2024_table.txt   (exportado de Enrichr)
    - KEGG_2026_table.txt                (exportado de Enrichr)
    - GO_Biological_Process_2025_table__1_.txt  (exportado de Enrichr)

SALIDAS:
    - figures/class_genes/figura1_dendrograma.png   : árbol del clustering
    - figures/class_genes/figura2_heatmap.png       : matriz genes x vías visualizada
    - figures/class_genes/figura3_barplot.png       : tamaño de cada cluster
    - results/clasificacion_final.xlsx              : clasificación de los 171 genes
    - results/matriz_genes_vias.csv                 : matriz binaria completa

CÓMO EJECUTAR (desde la raíz del proyecto HCC_singlecell_project):
    python scripts/clustering_171genes.py
=============================================================================
"""

# =============================================================================
# PASO 0 — IMPORTAR LIBRERÍAS
# =============================================================================
# pandas: leer y manipular tablas
# numpy: operaciones matemáticas con matrices
# scipy: clustering jerárquico y distancias
# matplotlib / seaborn: gráficos
# openpyxl: exportar a Excel
# collections: estructuras de datos auxiliares

import pandas as pd
import numpy as np
from collections import defaultdict, Counter
from scipy.cluster.hierarchy import linkage, fcluster, dendrogram
from scipy.spatial.distance import pdist, squareform
from scipy.stats import hypergeom
from statsmodels.stats.multitest import multipletests
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import os
from pathlib import Path

print("=" * 60)
print("CLUSTERING JERÁRQUICO DE GENES — TFM")
print("=" * 60)

# =============================================================================
# RUTAS DEL PROYECTO
# =============================================================================
# El script está en scripts/ y se ejecuta desde la raíz del proyecto.
#
# Estructura esperada:
#   HCC_singlecell_project/
#     scripts/          ← este script va aquí
#     data/             ← pon aquí los 3 archivos .txt de Enrichr
#     figures/
#       class_genes/    ← aquí se guardarán las figuras automáticamente
#     results/          ← aquí se guardará el Excel y el CSV

SCRIPT_DIR = Path(__file__).resolve().parent   # .../scripts/
BASE_DIR   = SCRIPT_DIR.parent                 # .../HCC_singlecell_project/

DATA_DIR = BASE_DIR / 'data' / 'enrichr_171genes'
FIG_DIR  = BASE_DIR / 'figures' / 'class_genes'
RES_DIR  = BASE_DIR / 'results'

# Crear carpetas de salida si no existen
FIG_DIR.mkdir(parents=True, exist_ok=True)
RES_DIR.mkdir(parents=True, exist_ok=True)

print(f"\nRutas del proyecto:")
print(f"  Raíz:    {BASE_DIR}")
print(f"  Datos:   {DATA_DIR}")
print(f"  Figuras: {FIG_DIR}")
print(f"  Results: {RES_DIR}")

# =============================================================================
# PASO 1 — DEFINIR LA LISTA DE GENES
# =============================================================================
# Estos son los 171 genes de tu TFM.

GENES = [
    'COL11A1','COL12A1','COL16A1','COL1A1','COL1A2','COL5A1','COL5A2','COL5A3',
    'COL6A1','COL6A2','COL6A3','COL9A2','BGN','ASPN','FMOD','VCAN','SPARC','POSTN',
    'TNC','THBS2','THBS3','LOX','LOXL1','LOXL3','PLOD1','P4HA3','FBN1','FBLN2',
    'EFEMP2','EMILIN2','MFAP4','MFAP5','CPXM1','CTHRC1','TIMP1','MMP9','MMP19',
    'MMP23B','LTBP2','SULF1','SULF2','CCN5','IGFBP3','ACVRL1','ANGPTL2','CCBE1',
    'PDGFRL','ENPP1','WNT16','SLIT2','SLIT3','PLXNA4','SOD3','CCL7','IL10','IL1RL1',
    'IL4R','ITGAM','LSP1','LILRB1','LILRA1','LILRB2','LILRA2','MILR1','TREM2',
    'CSF2RB','CLEC4D','CLEC11A','HMOX1','IFITM10','SRGN','SCARF2','OSTF1','HLA-A',
    'HLA-B','HLA-C','HLA-F','VSIR','TNFSF14','TNFRSF10B','TNFRSF10C','TNFRSF10D',
    'RUNX1','C3AR1','C5AR1','ADAM12','ADAM8','ADGRD1','ADRA2A','AGBL2','AHNAK2',
    'ALOX5AP','AQP3','ARL11','ARSJ','B3GNT3','BAALC','BHLHE22','CD248','CD276',
    'CHODL','CLDN10','CLIP3','CMTM3','CNDP2','CSGALNACT1','CTSH','CTSK','DIO3',
    'ECRG4','EEF1AKMT3','F2RL2','FKBP14','FKBP1B','GADD45B','GALNT10','GALNT9',
    'GPR153','GPX7','GRIA3','IRX1','ITGA11','KCNJ15','KCNK13','KCTD11','KDELR3',
    'LAYN','LXN','M1AP','MS4A14','MS4A4A','MS4A4E','MS4A6A','MS4A6E','NDRG4',
    'NSG1','PAK3','PCSK5','PHETA2','PLEKHG4','PRAF2','PRUNE2','PTPN5','RAP2A',
    'RASL12','RCN3','RFLNA','RIPK3','SELENOM','SEPTIN5','SFRP1','SFRP4','SGK1',
    'SHOX2','SLC37A2','SLC5A7','SLC6A7','SNX7','SORCS2','SPON1','SRPX2','STAC2',
    'SYT17','TCFL5','THY1','TMEM169','TRAM1L1','TSPAN18','TTC12','UBTD1','VSTM5'
]

print(f"\nPASO 1: {len(GENES)} genes cargados")

# =============================================================================
# PASO 2 — LEER LOS ARCHIVOS DE ENRICHR
# =============================================================================
# Enrichr nos da para cada vía:
#   - Term: nombre de la vía
#   - Genes: qué genes de nuestra lista están en esa vía
#   - Adjusted P-value: FDR (corrección de Benjamini-Hochberg)
#   - Odds Ratio: cuánto más representados están nuestros genes vs azar
#
# IMPORTANTE: Leemos TODAS las vías, no solo FDR<0.05.
# Esto es lo mismo que hace DAVID con su "flat annotation matrix".

reactome = pd.read_csv(DATA_DIR / 'Reactome_Pathways_2024_table.txt', sep='\t')
kegg     = pd.read_csv(DATA_DIR / 'KEGG_2026_table.txt', sep='\t')
gobp     = pd.read_csv(DATA_DIR / 'GO_Biological_Process_2025_table__1_.txt', sep='\t')

print(f"\nPASO 2: Archivos leídos")
print(f"  Reactome: {len(reactome)} vías totales")
print(f"  KEGG:     {len(kegg)} vías totales")
print(f"  GO BP:    {len(gobp)} vías totales")

# =============================================================================
# PASO 3 — CONSTRUIR EL MAPA GEN → VÍAS
# =============================================================================
# Para cada gen, guardamos en qué vías aparece.
# También guardamos info de cada vía (FDR, odds ratio) para usarla después
# en la caracterización de los clusters.
#
# Resultado: gene_terms[gen] = {via1, via2, via3, ...}

gene_terms  = defaultdict(set)   # gen -> conjunto de IDs de vías
term_info   = {}                  # via_id -> info de la vía (fdr, nombre, db)

for db_name, df in [('Reactome', reactome), ('KEGG', kegg), ('GO_BP', gobp)]:
    for _, row in df.iterrows():
        if pd.notna(row['Genes']) and pd.notna(row['Term']):
            term_id = f"{db_name}:{row['Term']}"
            genes_en_via = [g.strip() for g in str(row['Genes']).split(';')
                           if g.strip() in GENES]
            if genes_en_via:
                term_info[term_id] = {
                    'nombre': row['Term'],
                    'db': db_name,
                    'fdr': row['Adjusted P-value'],
                    'pval': row['P-value'],
                    'odds': row['Odds Ratio'],
                    'genes': genes_en_via
                }
                for gene in genes_en_via:
                    gene_terms[gene].add(term_id)

# Separar genes con y sin anotación
genes_clasificables = [g for g in GENES if g in gene_terms]
genes_sin_anotacion = [g for g in GENES if g not in gene_terms]

print(f"\nPASO 3: Mapa gen → vías construido")
print(f"  Vías con al menos 1 gen de la lista: {len(term_info)}")
print(f"  Genes con al menos 1 vía: {len(genes_clasificables)}")
print(f"  Genes sin ninguna vía en Enrichr: {len(genes_sin_anotacion)}")
if genes_sin_anotacion:
    print(f"  → Estos requieren anotación manual en GeneCards:")
    print(f"    {', '.join(genes_sin_anotacion)}")

# =============================================================================
# PASO 4 — CONSTRUIR LA MATRIZ BINARIA genes x vías
# =============================================================================
# Esta es la pieza central del método, igual que en DAVID.
#
# La matriz tiene:
#   - Filas: genes clasificables (151)
#   - Columnas: todas las vías con al menos 1 gen (1912)
#   - Valor: 1 si el gen está en esa vía, 0 si no
#
# Ejemplo de fila para COL1A1:
#   [1, 1, 1, 0, 0, 0, 1, 0, ...]  <- está en muchas vías de ECM
#
# Ejemplo de fila para HLA-A:
#   [0, 0, 0, 1, 1, 1, 0, 1, ...]  <- está en vías del sistema inmune
#
# COL1A1 y HLA-A tienen perfiles MUY distintos → distancia Jaccard alta → 
# el clustering los separa en grupos diferentes. ✓

term_ids = list(term_info.keys())  # lista de todas las vías (columnas)

matrix = np.zeros((len(genes_clasificables), len(term_ids)), dtype=np.int8)
for i, gene in enumerate(genes_clasificables):
    for j, term_id in enumerate(term_ids):
        if term_id in gene_terms[gene]:
            matrix[i, j] = 1

print(f"\nPASO 4: Matriz binaria construida")
print(f"  Dimensiones: {matrix.shape[0]} genes x {matrix.shape[1]} vías")
print(f"  Densidad: {matrix.sum() / matrix.size * 100:.1f}% de celdas con valor 1")

# =============================================================================
# PASO 5 — CALCULAR DISTANCIA JACCARD ENTRE GENES
# =============================================================================
# La distancia Jaccard entre dos genes A y B mide cuán DIFERENTES son:
#
#   Jaccard(A, B) = 1 - |vías compartidas| / |vías en A o en B|
#
# Ejemplos:
#   - COL1A1 y COL1A2 comparten casi todas sus vías → Jaccard ≈ 0 (muy similares)
#   - COL1A1 y HLA-A no comparten casi ninguna vía  → Jaccard ≈ 1 (muy distintos)
#
# pdist() calcula todas las distancias por pares (N*(N-1)/2 valores)
# Es la métrica estándar para vectores binarios.
# Referencia: SimplifyEnrichment (Gu & Hübschmann, 2023, Genomics Proteomics Bioinformatics)

print(f"\nPASO 5: Calculando distancias Jaccard entre {len(genes_clasificables)} genes...")
dist_condensed = pdist(matrix, metric='jaccard')
print(f"  Calculadas {len(dist_condensed):,} distancias por pares")

# =============================================================================
# PASO 6 — CLUSTERING JERÁRQUICO CON MÉTODO WARD
# =============================================================================
# El clustering jerárquico construye un árbol (dendrograma) fusionando genes
# de más similares a menos similares:
#
# Inicio: 151 grupos (cada gen es su propio grupo)
# Paso 1: une los 2 genes más similares (Jaccard más bajo) → 150 grupos
# Paso 2: une el siguiente par más similar → 149 grupos
# ...
# Final: 1 grupo con todos los genes
#
# MÉTODO WARD: en cada fusión, elige el par que minimiza el aumento de
# varianza interna de los grupos. Produce grupos compactos y bien separados.
# Es el método más usado en bioinformática para este tipo de análisis.
#
# linkage() devuelve la matriz Z que codifica el árbol completo.

print(f"\nPASO 6: Clustering jerárquico (método Ward)...")
Z = linkage(dist_condensed, method='ward')
print(f"  Árbol construido con {len(Z)} fusiones")

# =============================================================================
# PASO 7 — DETECCIÓN AUTOMÁTICA DE CLUSTERS POR ENRIQUECIMIENTO
# =============================================================================
# En lugar de cortar el árbol a una altura fija (como con el silhouette),
# recorremos el árbol de arriba a abajo y en cada nodo preguntamos:
# ¿los genes de este nodo tienen enriquecimiento funcional significativo?
#
# ALGORITMO (inspirado en funcExplorer, Kolberg et al. 2018 BMC Genomics):
#
#   1. Empezamos en la raíz del árbol (todos los genes juntos)
#   2. En cada nodo calculamos un test de enriquecimiento hipergeométrico
#      para las vías de nuestras bases de datos
#   3. Si hay al menos una vía con FDR < 0.05 → este nodo es un CLUSTER
#      No bajamos más por esta rama. Los genes de este nodo forman un grupo.
#   4. Si no hay enriquecimiento → bajamos a los dos nodos hijos y repetimos
#   5. Si llegamos a un gen individual sin enriquecimiento en ningún nodo
#      ancestral → ese gen queda SIN CLASIFICAR
#
# POR QUÉ ESTO ES MEJOR QUE K FIJO:
#   - Los clusters emergen donde hay señal biológica real
#   - No hay que justificar por qué K=8 y no K=7
#   - Cada cluster tiene garantizado enriquecimiento significativo
#   - Los genes sin señal funcional quedan honestamente sin clasificar
#
# TEST HIPERGEOMÉTRICO EN CADA NODO:
#   Para cada vía con M genes totales en la base de datos:
#   - N = total de genes clasificables (151)
#   - n = genes del nodo actual
#   - k = genes del nodo que están en esa vía
#   p-value = P(X >= k) donde X ~ Hipergeométrica(N, M, n)
#   Corrección FDR por Benjamini-Hochberg sobre todos los tests del nodo.
#
# Referencia: Kolberg et al. (2018) BMC Genomics 19:817
#             Rousseeuw (1987) J. Comput. Appl. Math. 20:53-65 (silhouette)

from sklearn.metrics import silhouette_score

# Convertir distancias a matriz cuadrada (necesario para silhouette)
dist_matrix_sq = squareform(dist_condensed)

N_TOTAL = len(genes_clasificables)   # universo = genes clasificables
FDR_THRESHOLD = 0.05                 # umbral de significación
MIN_CLUSTER_SIZE = 3                 # clusters con < 3 genes se ignoran

# --- Preparar información de vías para el test hipergeométrico ---
# Para cada vía necesitamos saber cuántos de nuestros genes clasificables
# están en ella (M = tamaño de la vía en nuestro universo)
term_genes_set = {}   # term_id -> set de genes clasificables en esa vía
for tid, info in term_info.items():
    genes_en_via = set(info.get('genes', []))
    genes_en_via_nuestros = genes_en_via & set(genes_clasificables)
    if len(genes_en_via_nuestros) >= 2:  # solo vías con >=2 de nuestros genes
        term_genes_set[tid] = genes_en_via_nuestros

term_ids_for_test = list(term_genes_set.keys())
print(f"\nPASO 7: Detección automática de clusters por enriquecimiento...")
print(f"  Universo: {N_TOTAL} genes clasificables")
print(f"  Vías candidatas para tests: {len(term_ids_for_test)}")
print(f"  Umbral FDR: {FDR_THRESHOLD}")

def test_enrichment_node(node_genes):
    """
    Calcula el enriquecimiento funcional para un conjunto de genes (nodo del árbol).
    Devuelve lista de (term_id, p_value, fdr) ordenada por FDR.
    Solo usa vías que tengan al menos 2 genes del nodo.
    """
    node_set = set(node_genes)
    n = len(node_set)
    if n < MIN_CLUSTER_SIZE:
        return []

    pvals = []
    tids = []
    for tid in term_ids_for_test:
        M = len(term_genes_set[tid])          # genes de la vía en nuestro universo
        k = len(node_set & term_genes_set[tid])  # genes del nodo en la vía
        if k < 2:
            continue
        # Test hipergeométrico: P(X >= k)
        # hypergeom.sf(k-1, N, M, n) = P(X >= k)
        pval = hypergeom.sf(k - 1, N_TOTAL, M, n)
        pvals.append(pval)
        tids.append(tid)

    if not pvals:
        return []

    # Corrección FDR Benjamini-Hochberg
    _, fdr_vals, _, _ = multipletests(pvals, method='fdr_bh')

    results = sorted(zip(tids, pvals, fdr_vals), key=lambda x: x[2])
    return results

def get_node_genes(node_id, Z, n_leaves):
    """
    Dado un nodo interno del árbol (Z), devuelve los índices de sus hojas (genes).
    Los nodos internos en Z tienen id >= n_leaves.
    Los nodos hoja tienen id < n_leaves.
    """
    if node_id < n_leaves:
        return [int(node_id)]
    left  = int(Z[int(node_id) - n_leaves, 0])
    right = int(Z[int(node_id) - n_leaves, 1])
    return get_node_genes(left, Z, n_leaves) + get_node_genes(right, Z, n_leaves)

# --- Recorrer el árbol de arriba a abajo ---
# El árbol tiene n_leaves hojas (genes) y n_leaves-1 nodos internos.
# La raíz es el nodo con id = 2*n_leaves - 2.
n_leaves = len(genes_clasificables)
root_id  = 2 * n_leaves - 2

cluster_genes   = {}   # cid -> lista de genes
cluster_enrich  = {}   # cid -> lista de (term_id, pval, fdr) del nodo
cluster_counter = [0]  # contador de clusters (lista para modificarlo en función recursiva)

def traverse(node_id):
    """
    Recorre el árbol recursivamente.
    Si el nodo tiene enriquecimiento significativo → es un cluster.
    Si no → baja a los dos hijos.
    """
    gene_indices = get_node_genes(node_id, Z, n_leaves)
    node_genes   = [genes_clasificables[i] for i in gene_indices]

    if len(node_genes) < MIN_CLUSTER_SIZE:
        # Nodo demasiado pequeño → sin clasificar
        return

    enrich_results = test_enrichment_node(node_genes)
    sig_results    = [(t, p, f) for t, p, f in enrich_results if f < FDR_THRESHOLD]

    if sig_results:
        # Este nodo tiene enriquecimiento significativo → es un cluster
        cluster_counter[0] += 1
        cid = cluster_counter[0]
        cluster_genes[cid]  = node_genes
        cluster_enrich[cid] = sig_results
    else:
        # No hay enriquecimiento → bajar a los hijos
        if node_id >= n_leaves:
            left  = int(Z[node_id - n_leaves, 0])
            right = int(Z[node_id - n_leaves, 1])
            traverse(left)
            traverse(right)

# Aumentar límite de recursión para árboles grandes
import sys
sys.setrecursionlimit(10000)

print(f"  Recorriendo árbol ({n_leaves} hojas, {n_leaves-1} nodos internos)...")
traverse(root_id)

# Genes sin clasificar = genes que no están en ningún cluster
genes_en_clusters = set()
for cg in cluster_genes.values():
    genes_en_clusters.update(cg)
genes_no_clasificados_arbol = [g for g in genes_clasificables
                                if g not in genes_en_clusters]

K = len(cluster_genes)
print(f"\n  Clusters detectados automáticamente: {K}")
print(f"  Genes clasificados: {len(genes_en_clusters)}")
print(f"  Genes clasificables sin cluster (sin enriquecimiento): {len(genes_no_clasificados_arbol)}")

# Construir vector de labels para silhouette y dendrograma
# (genes sin cluster → label 0)
cluster_labels = np.zeros(n_leaves, dtype=int)
for cid, cg in cluster_genes.items():
    for gene in cg:
        idx = genes_clasificables.index(gene)
        cluster_labels[idx] = cid

# --- Silhouette como validación (no para elegir K, sino para validar) ---
# Solo sobre genes que tienen cluster asignado
mask = cluster_labels > 0
if mask.sum() > 1 and len(set(cluster_labels[mask])) > 1:
    sil_val = silhouette_score(
        dist_matrix_sq[np.ix_(mask, mask)],
        cluster_labels[mask],
        metric='precomputed'
    )
    print(f"  Silhouette de validación: {sil_val:.4f}")
else:
    sil_val = None
    print(f"  Silhouette: no calculable (muy pocos clusters)")

# --- Mostrar resumen de cada cluster ---
print(f"\n  Resumen de clusters:")
for cid in sorted(cluster_genes.keys()):
    cg  = cluster_genes[cid]
    top = cluster_enrich[cid][0]  # vía más significativa
    print(f"  Cluster {cid} ({len(cg)} genes): "
          f"top vía = {term_info[top[0]]['nombre'][:50]} "
          f"(FDR={top[2]:.2e})")

# =============================================================================
# PASO 8 — CARACTERIZAR CADA CLUSTER CON SUS VÍAS MÁS SIGNIFICATIVAS
# =============================================================================
# En este método, los clusters ya tienen sus vías de enriquecimiento calculadas
# (se calcularon durante la detección en el paso 7).
#
# Aquí las organizamos para mostrar las top 10 por cluster:
#   - Ordenadas por FDR (menor = más significativo)
#   - Con el número de genes del cluster que están en cada vía

print(f"\nPASO 8: Organizando caracterización de clusters...")
cluster_top_terms = {}

for cid in sorted(cluster_genes.keys()):
    cg = cluster_genes[cid]
    # Las vías del cluster vienen del test de enriquecimiento del nodo
    # Añadimos el conteo de genes para cada vía
    top_terms_with_count = []
    for tid, pval, fdr in cluster_enrich[cid][:10]:
        count = len(set(cg) & term_genes_set.get(tid, set()))
        top_terms_with_count.append((tid, count))
    cluster_top_terms[cid] = top_terms_with_count

    print(f"\n  CLUSTER {cid} ({len(cg)} genes):")
    print(f"  Genes: {', '.join(cg[:8])}{'...' if len(cg)>8 else ''}")
    print(f"  Top 5 vías (test hipergeométrico en el nodo):")
    for tid, count in top_terms_with_count[:5]:
        info = term_info[tid]
        _, _, fdr = next(x for x in cluster_enrich[cid] if x[0] == tid)
        print(f"    [{info['db']}] {info['nombre'][:55]}")
        print(f"           {count}/{len(cg)} genes | FDR={fdr:.2e}")

# Figura silhouette — ahora muestra el valor de validación, no selección de K
fig, ax = plt.subplots(figsize=(7, 5))
if sil_val is not None:
    ax.bar(['Silhouette\nde validación'], [sil_val],
           color='#2E75B6', width=0.4, edgecolor='white')
    ax.axhline(y=0, color='gray', linestyle='--', linewidth=1)
    ax.text(0, sil_val + 0.005, f'{sil_val:.4f}',
            ha='center', va='bottom', fontsize=13, fontweight='bold')
    ax.set_ylim(-0.1, max(0.3, sil_val + 0.05))
    ax.set_ylabel('Coeficiente de silhouette', fontsize=12)
    ax.set_title(f'Validación del clustering — {K} clusters detectados automáticamente\n'
                 f'Método: enriquecimiento hipergeométrico por nodo (FDR<{FDR_THRESHOLD})\n'
                 f'Silhouette > 0 indica separación positiva entre clusters',
                 fontsize=11, fontweight='bold')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
else:
    ax.text(0.5, 0.5, 'Silhouette no calculable\n(menos de 2 clusters)',
            ha='center', va='center', transform=ax.transAxes, fontsize=12)

plt.tight_layout()
plt.savefig(FIG_DIR / 'figura0_validacion.png', dpi=150, bbox_inches='tight')
plt.close()
print(f"\n  Guardada: {FIG_DIR / 'figura0_validacion.png'}")

# =============================================================================
# PASO 9 — FIGURA 1: DENDROGRAMA
# =============================================================================
# El dendrograma muestra el árbol del clustering.
# - Eje X: genes (hojas del árbol)
# - Eje Y: distancia Ward a la que se fusionan
# - La línea horizontal de corte indica dónde dividimos en K grupos
# - Los genes del mismo color están en el mismo cluster
#
# En el TFM: esta figura justifica visualmente el número de clusters elegido.
# Los "saltos" en la altura del árbol indican dónde hay separación natural.

PALETTE = [
    '#C00000','#2E75B6','#375623','#7B3F00','#6A0DAD','#B8860B',
    '#006666','#CC5500','#8B0057','#1B5E20','#004D99','#7F3300',
    '#4A0080','#005555','#CC0044','#336600','#003366','#884400'
]
CLUSTER_COLORS = {i: PALETTE[(i-1) % len(PALETTE)] for i in range(1, 50)}
CLUSTER_COLORS[0] = '#888888'

print(f"\nPASO 9: Generando figura 1 (dendrograma)...")
fig, ax = plt.subplots(figsize=(22, 8))

# Línea de corte: altura donde se producen las K-1 últimas fusiones
cut_height = (Z[-(K-1), 2] + Z[-K, 2]) / 2

dend = dendrogram(
    Z,
    labels=genes_clasificables,
    leaf_rotation=90,
    leaf_font_size=6,
    color_threshold=cut_height,
    above_threshold_color='gray',
    ax=ax
)

# Línea de corte
ax.axhline(y=cut_height, color='black', linestyle='--', linewidth=1.5,
           label=f'Corte en k={K} clusters')

# Colorear etiquetas por cluster
leaf_order = dend['leaves']
for idx, leaf_pos in enumerate(leaf_order):
    gene = genes_clasificables[leaf_pos]
    cid = cluster_labels[leaf_pos]
    ax.get_xticklabels()[idx].set_color(CLUSTER_COLORS.get(cid, 'black'))

ax.set_title(f'Dendrograma — Clustering jerárquico de {len(genes_clasificables)} genes\n'
             f'Distancia Jaccard, método Ward, k={K} clusters',
             fontsize=14, fontweight='bold')
ax.set_ylabel('Distancia Ward', fontsize=11)
ax.set_xlabel('Genes', fontsize=10)

legend_patches = [mpatches.Patch(color=CLUSTER_COLORS[i], label=f'Cluster {i}')
                  for i in range(1, K+1)]
legend_patches.append(mpatches.Patch(color='none', label=''))
ax.legend(handles=legend_patches, loc='upper right', fontsize=9,
          title='Clusters', title_fontsize=10)

plt.tight_layout()
plt.savefig(FIG_DIR / 'figura1_dendrograma.png', dpi=150, bbox_inches='tight')
plt.close()
print(f"  Guardada: {FIG_DIR / 'figura1_dendrograma.png'}")

# =============================================================================
# PASO 10 — FIGURA 2: HEATMAP
# =============================================================================
# El heatmap muestra la matriz genes x vías en formato visual:
# - Rojo = gen presente en la vía (valor 1)
# - Gris claro = gen ausente (valor 0)
# - Genes ordenados según el dendrograma (mismo orden que figura 1)
# - Columnas: las 50 vías más significativas (FDR<0.05, ordenadas por FDR)
#
# En el TFM: esta figura muestra que los clusters tienen perfiles de vías
# distintos y coherentes. Los genes del mismo cluster tienen patrones similares.

print(f"\nPASO 10: Generando figura 2 (heatmap)...")

# Seleccionar top 50 vías significativas como columnas
sig_term_ids = sorted(
    [t for t in term_info if term_info[t]['fdr'] < 0.05],
    key=lambda t: term_info[t]['fdr']
)[:50]

# Reordenar genes según el dendrograma
leaf_order = dend['leaves']
genes_ordered = [genes_clasificables[i] for i in leaf_order]
clusters_ordered = [cluster_labels[i] for i in leaf_order]

# Submatriz: genes reordenados x top 50 vías
term_to_col = {t: j for j, t in enumerate(term_ids)}
hm_matrix = np.zeros((len(genes_ordered), len(sig_term_ids)), dtype=int)
for i, gene in enumerate(genes_ordered):
    for j, term_id in enumerate(sig_term_ids):
        if term_id in gene_terms.get(gene, set()):
            hm_matrix[i, j] = 1

# Etiquetas cortas para las vías
term_labels_short = []
for t in sig_term_ids:
    name = term_info[t]['nombre']
    name = name[:32] + '...' if len(name) > 32 else name
    term_labels_short.append(name)

fig, ax = plt.subplots(figsize=(24, max(12, len(genes_ordered) * 0.18)))

sns.heatmap(
    hm_matrix,
    xticklabels=term_labels_short,
    yticklabels=genes_ordered,
    cmap=['#F0F0F0', '#C00000'],
    linewidths=0.2,
    linecolor='white',
    cbar_kws={'label': '0 = ausente  |  1 = presente en la vía', 'shrink': 0.3},
    ax=ax
)

# Colorear etiquetas de genes por cluster
for i, (lbl, cid) in enumerate(zip(ax.get_yticklabels(), clusters_ordered)):
    lbl.set_color(CLUSTER_COLORS.get(cid, 'black'))
    lbl.set_fontsize(7)
    lbl.set_fontweight('bold')

ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right', fontsize=7)
ax.set_title(f'Heatmap — Presencia/ausencia de {len(genes_ordered)} genes en las 50 vías más significativas\n'
             f'Genes ordenados según clustering jerárquico | Colores = clusters',
             fontsize=13, fontweight='bold', pad=15)

legend_patches = [mpatches.Patch(color=CLUSTER_COLORS[i], label=f'Cluster {i}')
                  for i in range(1, K+1)]
ax.legend(handles=legend_patches, loc='upper left',
          bbox_to_anchor=(1.15, 1), fontsize=9, title='Clusters')

plt.tight_layout()
plt.savefig(FIG_DIR / 'figura2_heatmap.png', dpi=150, bbox_inches='tight')
plt.close()
print(f"  Guardada: {FIG_DIR / 'figura2_heatmap.png'}")

# =============================================================================
# PASO 11 — FIGURA 3: BARPLOT
# =============================================================================
# Resumen visual del número de genes en cada cluster.
# Simple pero útil para el TFM para mostrar el tamaño de cada grupo.

print(f"\nPASO 11: Generando figura 3 (barplot)...")
cluster_sizes = {cid: len(cluster_genes[cid]) for cid in range(1, K+1)}

fig, ax = plt.subplots(figsize=(9, 6))
bars = ax.bar(
    [f'Cluster {i}' for i in range(1, K+1)],
    [cluster_sizes[i] for i in range(1, K+1)],
    color=[CLUSTER_COLORS[i] for i in range(1, K+1)],
    edgecolor='white', linewidth=1.5, width=0.6
)
for bar, cid in zip(bars, range(1, K+1)):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
            str(cluster_sizes[cid]), ha='center', va='bottom',
            fontsize=13, fontweight='bold')

ax.set_ylabel('Número de genes', fontsize=12)
ax.set_title(
    f'Distribución de {len(genes_clasificables)} genes en {K} clusters funcionales\n'
    f'(+{len(genes_sin_anotacion)} genes sin anotación en Enrichr — requieren GeneCards)',
    fontsize=12, fontweight='bold'
)
ax.set_ylim(0, max(cluster_sizes.values()) + 10)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
plt.tight_layout()
plt.savefig(FIG_DIR / 'figura3_barplot.png', dpi=150, bbox_inches='tight')
plt.close()
print(f"  Guardada: {FIG_DIR / 'figura3_barplot.png'}")

# =============================================================================
# PASO 12 — EXPORTAR CLASIFICACIÓN A EXCEL
# =============================================================================
# Excel con 2 pestañas:
#   1. Clasificación final: cada gen con su cluster y las vías que lo definen
#   2. Genes sin anotación en Enrichr

print(f"\nPASO 12: Exportando clasificación a Excel...")

THIN = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
)
ALT = PatternFill('solid', start_color='F7F7F7')
FILL_PALETTE = [
    'FADADD','DAE8F5','D5E8D4','FDE8CE','EDE7F6','FFF9C4',
    'D6EEF0','FFE4CC','F5D0E8','DCEDC8','CCE5FF','FFD9B3',
    'E8D5F5','CCF2F2','FFB3C6','D4EDDA','CCE0FF','FFE0B2'
]
FILLS   = {i: FILL_PALETTE[(i-1) % len(FILL_PALETTE)] for i in range(1, 50)}
FILLS[0] = 'EEEEEE'
HCOLORS = {i: PALETTE[(i-1) % len(PALETTE)].replace('#','') for i in range(1, 50)}
HCOLORS[0] = '888888'

wb = Workbook()
ws = wb.active
ws.title = 'Clasificacion final'

ws.merge_cells('A1:F1')
ws['A1'] = (f'Clasificación funcional de {len(GENES)} genes — '
            f'Clustering jerárquico (Distancia Jaccard + Linkage Ward, k={K}) | '
            f'Datos: Enrichr Reactome + KEGG + GO BP (SIN filtro FDR para la matriz) | '
            f'Metodología inspirada en DAVID (Huang et al. 2007)')
ws['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
ws['A1'].fill = PatternFill('solid', start_color='1F1F1F')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 38

hdrs = ['Gen', 'Cluster', 'Nº vías (todas)', 'Nº vías significativas (FDR<0.05)',
        'Top 3 vías características del cluster (FDR<0.05)', 'Genes sin clasificar']
col_widths = [14, 10, 18, 30, 70, 20]
for col, (h, w) in enumerate(zip(hdrs, col_widths), 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', start_color='404040')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN
    ws.column_dimensions[c.column_letter].width = w
ws.row_dimensions[2].height = 22
ws.freeze_panes = 'A3'

row_idx = 3
for cid in range(1, K+1):
    cg = sorted(cluster_genes[cid])
    fill = PatternFill('solid', start_color=FILLS[cid])

    # Top 3 vías del cluster para mostrar en la tabla
    top3 = [term_info[t]['nombre'] for t, _ in cluster_top_terms[cid][:3]]
    top3_str = ' | '.join(top3)

    # Cabecera de sección
    ws.merge_cells(f'A{row_idx}:F{row_idx}')
    c = ws.cell(row=row_idx, column=1,
                value=f'Cluster {cid} — {len(cg)} genes')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', start_color=HCOLORS[cid])
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

    for i, gene in enumerate(cg):
        rf = fill if i % 2 == 0 else ALT
        n_all = len(gene_terms.get(gene, set()))
        n_sig = sum(1 for t in gene_terms.get(gene, set())
                   if term_info[t]['fdr'] < 0.05)
        vals = [gene, f'Cluster {cid}', n_all, n_sig, top3_str, '']
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = Font(name='Arial', size=10, bold=(col == 1))
            c.fill = rf; c.border = THIN
            c.alignment = Alignment(wrap_text=(col == 5), vertical='top')
        row_idx += 1

# Genes sin anotación
fill0 = PatternFill('solid', start_color=FILLS[0])
ws.merge_cells(f'A{row_idx}:F{row_idx}')
c = ws.cell(row=row_idx, column=1,
            value=f'Sin anotación en Enrichr — {len(genes_sin_anotacion)} genes (requieren GeneCards)')
c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
c.fill = PatternFill('solid', start_color=HCOLORS[0])
c.alignment = Alignment(vertical='center')
ws.row_dimensions[row_idx].height = 20
row_idx += 1

for i, gene in enumerate(sorted(genes_sin_anotacion)):
    rf = fill0 if i % 2 == 0 else ALT
    vals = [gene, 'Sin clasificar', 0, 0,
            'No aparece en ninguna vía de Enrichr', 'Consultar GeneCards / UniProt']
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.font = Font(name='Arial', size=10, bold=(col == 1),
                     color='888888' if col > 1 else '000000')
        c.fill = rf; c.border = THIN
        c.alignment = Alignment(vertical='top')
    row_idx += 1

# =============================================================================
# PESTAÑA 2 — CARACTERIZACIÓN DE CLUSTERS (top 10 vías por cluster)
# =============================================================================
# Esta pestaña es la clave para justificar el nombre que le das a cada cluster.
#
# Para cada cluster muestra las 10 vías significativas (FDR<0.05) en las que
# más genes del cluster aparecen. Con esta tabla puedes:
#   1. Ver qué función biológica predomina en cada cluster
#   2. Decidir el nombre del cluster basándote en las vías más frecuentes
#   3. Justificar ese nombre en el TFM con datos concretos (FDR, nº genes)
#
# Cómo leer la tabla:
#   - "Nº genes cluster en vía" → de los N genes del cluster, cuántos están en esa vía
#   - "% cobertura" → qué porcentaje del cluster está en esa vía
#   - "FDR" → significación estadística de esa vía en el análisis de enriquecimiento
#   - "Base de datos" → de dónde viene la vía (Reactome, KEGG o GO BP)
#
# Ejemplo de cómo usarla:
#   Si en el Cluster 1 ves que las top vías son todas de colágeno y ECM,
#   el nombre del cluster será "Síntesis y organización de la ECM".

ws2 = wb.create_sheet('Caracterizacion clusters')

ws2.merge_cells('A1:G1')
ws2['A1'] = ('Caracterización de clusters — Top 10 vías significativas (FDR<0.05) por cluster | '
             'Usa esta tabla para decidir y justificar el nombre de cada cluster')
ws2['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
ws2['A1'].fill = PatternFill('solid', start_color='1F1F1F')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws2.row_dimensions[1].height = 34

ws2.merge_cells('A2:G2')
ws2['A2'] = ('CÓMO USAR: Mira las vías más frecuentes de cada cluster (mayor % cobertura y menor FDR). '
             'El patrón de vías te indica la función biológica del grupo y te permite ponerle nombre.')
ws2['A2'].font = Font(name='Arial', italic=True, size=9, color='555555')
ws2['A2'].alignment = Alignment(horizontal='center', wrap_text=True)
ws2.row_dimensions[2].height = 22

hdrs2 = ['Cluster', 'Rango', 'Vía / Término', 'Base de datos',
         'Nº genes cluster en vía', '% cobertura del cluster', 'FDR (Adjusted P-value)']
col_widths2 = [12, 8, 55, 14, 24, 24, 22]
for col, (h, w) in enumerate(zip(hdrs2, col_widths2), 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', start_color='404040')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN
    ws2.column_dimensions[c.column_letter].width = w
ws2.row_dimensions[3].height = 22
ws2.freeze_panes = 'A4'

row_idx2 = 4
for cid in range(1, K+1):
    cg = cluster_genes[cid]
    n_genes_cluster = len(cg)
    fill = PatternFill('solid', start_color=FILLS[cid])

    # Cabecera de sección del cluster
    ws2.merge_cells(f'A{row_idx2}:G{row_idx2}')
    c = ws2.cell(row=row_idx2, column=1,
                 value=f'Cluster {cid} — {n_genes_cluster} genes | '
                       f'Genes: {", ".join(sorted(cg)[:10])}{"..." if len(cg)>10 else ""}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', start_color=HCOLORS[cid])
    c.alignment = Alignment(vertical='center', wrap_text=True)
    ws2.row_dimensions[row_idx2].height = 20
    row_idx2 += 1

    # Top 10 vías del cluster
    for rank, (term_id, count) in enumerate(cluster_top_terms[cid][:10], 1):
        info = term_info[term_id]
        pct = round(count / n_genes_cluster * 100, 1)
        rf = fill if rank % 2 == 0 else ALT

        vals = [
            f'Cluster {cid}',
            rank,
            info['nombre'],
            info['db'],
            f'{count} / {n_genes_cluster}',
            f'{pct}%',
            info['fdr']
        ]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row_idx2, column=col, value=val)
            c.font = Font(name='Arial', size=10,
                         bold=(rank == 1))  # negrita para la vía más frecuente
            c.fill = rf
            c.border = THIN
            c.alignment = Alignment(wrap_text=(col == 3), vertical='top',
                                   horizontal='center' if col != 3 else 'left')
            if col == 7:
                c.number_format = '0.00E+00'
            # Resaltar la vía #1 de cada cluster con borde más grueso
            if rank == 1:
                c.font = Font(name='Arial', size=10, bold=True,
                             color=HCOLORS[cid])
        row_idx2 += 1

    # Fila en blanco entre clusters
    row_idx2 += 1

# =============================================================================
# PESTAÑA 3 — SILHOUETTE
# =============================================================================
# Tabla con los valores de silhouette para cada K probado.
# Útil para incluir en el TFM como evidencia estadística de la elección de K.

ws3 = wb.create_sheet('Validacion clustering')

ws3.merge_cells('A1:D1')
ws3['A1'] = ('Validación del clustering por enriquecimiento automático | '
             'Método: test hipergeométrico por nodo + corrección FDR Benjamini-Hochberg | '
             'Referencia: Kolberg et al. (2018) BMC Genomics 19:817')
ws3['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
ws3['A1'].fill = PatternFill('solid', start_color='1F1F1F')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws3.row_dimensions[1].height = 40

ws3.merge_cells('A2:D2')
ws3['A2'] = ('En este método NO se elige K. Los clusters emergen automáticamente donde '
             'hay enriquecimiento significativo (FDR<0.05) en el árbol. '
             'El silhouette se usa como VALIDACIÓN, no para elegir K.')
ws3['A2'].font = Font(name='Arial', italic=True, size=9, color='993300')
ws3['A2'].alignment = Alignment(horizontal='center', wrap_text=True)
ws3.row_dimensions[2].height = 28

# Summary stats
stats = [
    ('Método de detección de clusters', 'Enriquecimiento hipergeométrico por nodo del dendrograma'),
    ('Corrección múltiples tests', 'Benjamini-Hochberg (FDR)'),
    ('Umbral FDR', str(FDR_THRESHOLD)),
    ('Tamaño mínimo de cluster', str(MIN_CLUSTER_SIZE) + ' genes'),
    ('Universo de genes', str(N_TOTAL) + ' genes clasificables'),
    ('Vías candidatas testadas', str(len(term_ids_for_test))),
    ('Clusters detectados', str(K)),
    ('Genes clasificados', str(len(genes_en_clusters))),
    ('Genes sin cluster (sin enriquecimiento)', str(len(genes_no_clasificados_arbol))),
    ('Silhouette de validación', f'{sil_val:.4f}' if sil_val else 'No calculable'),
    ('Referencia metodológica', 'Kolberg et al. (2018) BMC Genomics 19:817 (funcExplorer)'),
]

for col, h in enumerate(['Parámetro', 'Valor'], 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', start_color='404040')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN
ws3.column_dimensions['A'].width = 42
ws3.column_dimensions['B'].width = 55
ws3.row_dimensions[3].height = 20

for i, (param, val) in enumerate(stats, 4):
    rf = PatternFill('solid', start_color='F7F7F7') if i%2==0 else PatternFill('solid', start_color='FFFFFF')
    for col, v in enumerate([param, val], 1):
        c = ws3.cell(row=i, column=col, value=v)
        c.font = Font(name='Arial', size=10, bold=(col==1))
        c.fill = rf
        c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(RES_DIR / 'clasificacion_final.xlsx')
print(f"  Guardada: {RES_DIR / 'clasificacion_final.xlsx'}")

print("\n" + "=" * 60)
print("PROCESO COMPLETADO")
print("=" * 60)
print(f"  Método: detección automática por enriquecimiento (funcExplorer)")
print(f"  Clusters detectados: {K}")
print(f"  Genes clasificados:  {len(genes_en_clusters)} / {N_TOTAL}")
print(f"  Genes sin cluster:   {len(genes_no_clasificados_arbol)} (sin enriquecimiento en ningún nodo)")
print(f"  Genes sin Enrichr:   {len(genes_sin_anotacion)} (no aparecen en ninguna vía)")
if sil_val:
    print(f"  Silhouette validación: {sil_val:.4f}")
print(f"  Figuras en:  {FIG_DIR}")
print(f"    - figura0_validacion.png  ← silhouette de validación")
print(f"    - figura1_dendrograma.png ← árbol del clustering")
print(f"    - figura2_heatmap.png     ← matriz genes x vías")
print(f"    - figura3_barplot.png     ← tamaño de clusters")
print(f"  Excel en:    {RES_DIR / 'clasificacion_final.xlsx'}")
print("=" * 60)
